"""
主入口 — 商品库语义清洗去重
"""
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from translator import SemanticTranslator, ProductRow
from rules import RuleEngine
from report import print_summary


TEMPLATE_PATH = r"d:\Users\weis\Desktop\编码整理\商品库整理.xlsx"
OUTPUT_PATH = r"d:\Users\weis\Desktop\编码整理\商品库整理_结果.xlsx"


def load_rows(ws) -> list[ProductRow]:
    """从Excel加载并翻译所有行"""
    rows = []
    translator = SemanticTranslator()
    for r in range(2, ws.max_row + 1):
        raw_name = str(ws.cell(r, 2).value or '').strip()
        raw_unit = str(ws.cell(r, 3).value or '').strip()
        raw_desc = str(ws.cell(r, 4).value or '').strip()
        raw_cat = str(ws.cell(r, 5).value or '').strip()
        spuid = str(ws.cell(r, 1).value or '').strip()

        product = ProductRow(
            spuid=spuid,
            name_meaning=translator.translate_name(raw_name, raw_cat),
            unit=raw_unit,
            desc_meaning=translator.translate_desc(raw_desc, raw_unit),
            category=raw_cat,
        )
        rows.append(product)
    return rows


def write_results(ws, rows: list[ProductRow]):
    """写入列6-9"""
    for idx, row in enumerate(rows):
        r = idx + 2
        ws.cell(r, 6, value=row.group_id)
        ws.cell(r, 7, value=row.group_desc)
        ws.cell(r, 8, value=row.suggestion)
        ws.cell(r, 9, value=row.anomaly_class)


def main():
    print("=== 商品库语义清洗去重 ===")
    print(f"输入: {TEMPLATE_PATH}")
    print(f"输出: {OUTPUT_PATH}\n")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Sheet (2)"]
    total = ws.max_row - 1

    print(f"读入 {total} 行...")
    rows = load_rows(ws)
    print(f"翻译完成。\n")

    engine = RuleEngine(rows, SemanticTranslator())
    engine.run_all()
    summary = engine.summary()
    print_summary(summary, engine.conflicts_cat, engine.conflicts_spec, total)

    write_results(ws, rows)
    wb.save(OUTPUT_PATH)
    print(f"\n已写入: {OUTPUT_PATH}")

    final_count = sum(1 for row in rows if row.anomaly_class is not None)
    assert final_count == total, f"行数校验失败: {final_count} != {total}"
    print(f"行数校验通过: {final_count} = {total} ✓")


if __name__ == "__main__":
    main()
