"""
基准库索引构建 — 用新翻译器重建12440条索引
输入: 商品库整理_结果.xlsx
输出: index.json
"""
import json, sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from collections import defaultdict
from translator import SemanticTranslator, resolve_spec

SRC = r"d:\Users\weis\Desktop\编码整理\商品库整理_结果.xlsx"
DST = r"d:\Users\weis\Desktop\编码整理\index.json"


def run():
    print("=== 基准库索引构建 ===")
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Sheet (2)"]
    t = SemanticTranslator()

    index = defaultdict(list)
    total = 0
    errors = 0

    for r in range(2, ws.max_row + 1):
        spuid = str(ws.cell(r, 1).value or '').strip()
        raw_name = str(ws.cell(r, 2).value or '').strip()
        unit = str(ws.cell(r, 3).value or '').strip()
        raw_desc = str(ws.cell(r, 4).value or '').strip()
        category = str(ws.cell(r, 5).value or '').strip()
        hit_count = int(ws.cell(r, 11).value or 0)

        try:
            nm = t.translate_name(raw_name, category)
            dm = t.translate_desc(raw_desc, unit)
            spec = resolve_spec(nm, dm, has_proc=False)

            idx_row = {
                "spuid": spuid, "core": nm.core,
                "brand": nm.brand, "spec_core": spec.spec_core,
                "spec_conflict": spec.conflict,
                "has_processing": spec.has_processing,
                "unit": unit, "raw_name": raw_name, "category": category,
                "has_chaoma": nm.has_chaoma, "has_status": nm.has_status,
                "hit_count": hit_count,
            }
            index[nm.core].append(idx_row)
            total += 1
        except Exception as e:
            print(f"  ERROR row {r} {spuid}: {e}")
            errors += 1

    wb.close()

    index_dict = dict(index)
    with open(DST, 'w', encoding='utf-8') as f:
        json.dump(index_dict, f, ensure_ascii=False)

    print(f"翻译: {total}条, 错误: {errors}")
    print(f"索引core数: {len(index_dict)}")
    print(f"文件大小: {os.path.getsize(DST)/1024/1024:.1f}MB")
    print(f"已保存: {DST}")


if __name__ == "__main__":
    run()
