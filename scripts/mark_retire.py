"""
下架标记脚本：对完全重复和抄码名重复组，逐组判断保留/下架
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from collections import defaultdict


SRC = r"d:\Users\weis\Desktop\编码整理\商品库整理_结果.xlsx"
DST = r"d:\Users\weis\Desktop\编码整理\商品库整理_结果_下架_保留.xlsx"


def run():
    print("=== 下架标记 ===\n")
    wb = openpyxl.load_workbook(SRC)
    ws = wb["Sheet (2)"]

    # 1. 收集组
    groups = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        ac = str(ws.cell(r, 9).value or '').strip()
        if ac not in ("完全重复", "抄码名重复"):
            continue
        gid = str(ws.cell(r, 6).value or '').strip()
        if not gid:
            continue
        groups[gid].append(r)

    print(f"待处理组数: {len(groups)}")

    # 2. 确保表头
    ws.cell(1, 10, value="是否下架/保留")
    ws.cell(1, 13, value="替代编码")

    keep_count = 0
    retire_count = 0
    total_spus = sum(len(rows) for rows in groups.values())

    for gid, rows in groups.items():
        items = []
        for r in rows:
            name = str(ws.cell(r, 2).value or '').strip()
            hits = int(ws.cell(r, 11).value or 0)  # 列11: 命中次数
            spec = str(ws.cell(r, 4).value or '').strip()
            items.append({
                'row': r,
                'spuid': str(ws.cell(r, 1).value or '').strip(),
                'name': name,
                'hits': hits,
                'has_chaoma': '抄码' in name,
                'has_spec': bool(spec and spec not in ('', '抄码', '无', '散称', '散装')),
            })

        all_zero = all(it['hits'] == 0 for it in items)

        if all_zero:
            retain = pick_best(items)
        else:
            has_hits = [it for it in items if it['hits'] > 0]
            retain = max(has_hits, key=lambda it: (it['hits'], not it['has_chaoma'], it['has_spec']))

        retired = [it for it in items if it['spuid'] != retain['spuid']]
        retired_spuids = [it['spuid'] for it in retired]

        # 保留行: 替代编码 = 被它替代的编码列表
        ws.cell(retain['row'], 10, value="保留")
        ws.cell(retain['row'], 13, value=", ".join(retired_spuids) if retired_spuids else "")
        keep_count += 1

        # 下架行: 替代编码 = 保留的SPUID
        for alt in retired:
            ws.cell(alt['row'], 10, value="下架")
            ws.cell(alt['row'], 13, value=retain['spuid'])
            retire_count += 1

    print(f"保留: {keep_count}")
    print(f"下架: {retire_count}")
    print(f"合计: {keep_count + retire_count} / {total_spus}")
    assert keep_count + retire_count == total_spus

    wb.save(DST)
    print(f"\n已保存: {DST}")

    # 打印样例
    print("\n--- 样例 ---")
    for i, gid in enumerate(sorted(groups.keys())[:5]):
        for r in groups[gid]:
            action = ws.cell(r, 10).value
            spuid = ws.cell(r, 1).value
            name = str(ws.cell(r, 2).value or '')[:25]
            hits = ws.cell(r, 11).value
            replace = str(ws.cell(r, 13).value or '')
            print(f"  {gid} | {action} | {spuid} | {name} | 命中{hits} | 替代={replace}")
        if i < 4:
            print()


def pick_best(items):
    """全部0命中时选保留"""
    return max(items, key=lambda it: (
        not it['has_chaoma'],
        it['has_spec'],
        -len(it['name']),
    ))


if __name__ == "__main__":
    run()
