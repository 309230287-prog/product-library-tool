"""
交叉匹配脚本：将销售订单透视数据匹配到一期结果表
输入: 订单明细.xlsx (透视sheet), 商品库整理_结果.xlsx
输出: 商品库整理_结果.xlsx (填充第10、11列)
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from collections import defaultdict


ORDER_PATH = r"d:\Users\weis\Desktop\编码整理\订单明细.xlsx"
RESULT_PATH = r"d:\Users\weis\Desktop\编码整理\商品库整理_结果.xlsx"


def load_pivot(path: str) -> dict[str, tuple[int, list[str]]]:
    """读取透视表，返回 {SPUID: (总命中次数, [客户列表])}"""
    print("读取透视表...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["透视"]

    spuid_data: dict[str, tuple[int, list[str]]] = {}
    current_spuid = None

    row_count = 0
    for row in ws.iter_rows(min_row=5, values_only=True):  # row 4 is header
        spuid = str(row[0] or '').strip()
        customer = str(row[1] or '').strip()
        qty = row[5] or 0

        if not spuid:
            spuid = current_spuid
        else:
            current_spuid = spuid

        if not spuid or not customer:
            continue

        qty = int(qty)
        if spuid in spuid_data:
            prev_qty, prev_custs = spuid_data[spuid]
            spuid_data[spuid] = (prev_qty + qty, prev_custs + [customer])
        else:
            spuid_data[spuid] = (qty, [customer])

        row_count += 1

    wb.close()

    # 去重客户并排序
    result = {}
    for spuid, (total_qty, custs) in spuid_data.items():
        unique_custs = sorted(set(custs))
        result[spuid] = (total_qty, unique_custs)

    print(f"透视表读入 {row_count} 行, 去重后 {len(result)} 个SPUID")

    # 打印前几条样例
    for i, (spuid, (qty, custs)) in enumerate(sorted(result.items())[:5]):
        print(f"  {spuid}: {qty}次, {len(custs)}个客户")
        for c in custs[:3]:
            print(f"    - {c}")
        if len(custs) > 3:
            print(f"    ... 等{len(custs)}个客户")

    return result


def fill_results(pivot_data: dict, path: str):
    """将透视数据填入结果表第10、11列"""
    print(f"\n读取结果表: {path}")
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet (2)"]

    # 确保表头
    ws.cell(1, 11, value="订单命中次数")
    ws.cell(1, 12, value="订单命中客户")

    matched = 0
    unmatched = 0
    total = ws.max_row - 1

    for r in range(2, ws.max_row + 1):
        spuid = str(ws.cell(r, 1).value or '').strip()
        if spuid in pivot_data:
            qty, custs = pivot_data[spuid]
            ws.cell(r, 11, value=qty)
            ws.cell(r, 12, value="\n".join(custs))
            matched += 1
        else:
            ws.cell(r, 11, value=0)
            ws.cell(r, 12, value="")
            unmatched += 1

    print(f"匹配: {matched}, 未匹配: {unmatched}, 总计: {total}")

    wb.save(path)
    print(f"已保存: {path}")


def print_summary():
    """统计重复项的使用情况"""
    print("\n=== 重复项使用统计 ===")
    wb = openpyxl.load_workbook(RESULT_PATH, data_only=True)
    ws = wb["Sheet (2)"]

    # 按异常分类统计
    from collections import Counter
    stats: dict[str, dict] = {}

    for r in range(2, ws.max_row + 1):
        ac = str(ws.cell(r, 9).value or '').strip()
        if ac == "正常":
            continue
        hits = int(ws.cell(r, 11).value or 0)
        custs_str = str(ws.cell(r, 12).value or '')
        cust_count = len([c for c in custs_str.split("\n") if c.strip()]) if custs_str else 0

        if ac not in stats:
            stats[ac] = {"total": 0, "with_hits": 0, "total_hits": 0, "max_hits": 0}
        s = stats[ac]
        s["total"] += 1
        if hits > 0:
            s["with_hits"] += 1
        s["total_hits"] += hits
        s["max_hits"] = max(s["max_hits"], hits)

    print(f"{'分类':<12} {'总条数':<8} {'有使用':<8} {'总次数':<10} {'最多':<8}")
    print("-" * 50)
    for ac, s in sorted(stats.items()):
        print(f"{ac:<12} {s['total']:<8} {s['with_hits']:<8} {s['total_hits']:<10} {s['max_hits']:<8}")

    wb.close()


if __name__ == "__main__":
    print("=== 交叉匹配：订单数据 → 商品库结果表 ===\n")
    pivot_data = load_pivot(ORDER_PATH)
    fill_results(pivot_data, RESULT_PATH)
    print_summary()
    print("\n完成。")
