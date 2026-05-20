"""
新品查重主入口 — 读取新品新增接龙表，匹配后反写系统名称和编码
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from translator import SemanticTranslator
from matcher import load_index, check_one

SRC = r"d:\Users\weis\Desktop\编码整理\新品新增接龙表.xlsx"
DST = r"d:\Users\weis\Desktop\编码整理\新品新增接龙表_查重结果.xlsx"
INDEX_PATH = r"d:\Users\weis\Desktop\编码整理\index.json"


def run():
    print("=== 新品查重 ===\n")
    translator = SemanticTranslator()
    index = load_index(INDEX_PATH)

    wb = openpyxl.load_workbook(SRC)
    ws = wb["工作表1"]

    COL_BRAND, COL_NAME, COL_SPEC = 2, 3, 4
    COL_CATEGORY, COL_UNIT, COL_REMARK = 5, 6, 7
    COL_SYSNAME, COL_SYSCODE = 11, 12
    COL_RESULT, COL_DETAIL = 13, 14  # 新增：匹配结果+详情

    processed = matched = skipped = 0

    for r in range(2, ws.max_row + 1):
        sys_code = str(ws.cell(r, COL_SYSCODE).value or '').strip()
        if sys_code and sys_code != 'None':
            skipped += 1
            continue

        brand_raw = str(ws.cell(r, COL_BRAND).value or '').strip()
        if not brand_raw or brand_raw == 'None':
            brand = None
        elif brand_raw == '无要求':
            brand = '__SKIP__'  # 特殊标记：跳过品牌比对
        else:
            brand = brand_raw
        name = str(ws.cell(r, COL_NAME).value or '').strip()
        spec_desc = str(ws.cell(r, COL_SPEC).value or '').strip()
        category = str(ws.cell(r, COL_CATEGORY).value or '').strip()
        unit = str(ws.cell(r, COL_UNIT).value or '').strip()
        remark = str(ws.cell(r, COL_REMARK).value or '').strip()

        if not name:
            continue

        result = check_one(translator, name, brand, spec_desc, unit, category, remark, index)

        # 写结果列（所有结果类型都写）
        ws.cell(r, COL_RESULT, value=result.result)
        ws.cell(r, COL_DETAIL, value=result.detail)
        if result.result == "建议复用":
            best = next((c for c in result.candidates if c.spuid == result.suggested_spuid), None)
            best_name = best.raw_name if best else result.suggested_spuid
            ws.cell(r, COL_SYSNAME, value=best_name)
            ws.cell(r, COL_SYSCODE, value=result.suggested_spuid)
            matched += 1

        processed += 1
        if processed % 10 == 0:
            print(f"  已处理 {processed}, 匹配 {matched}")

    wb.save(DST)
    print(f"\n跳过(已有编码): {skipped}")
    print(f"处理: {processed}条, 复用: {matched}条, 待定: {processed - matched}条")
    print(f"已保存: {DST}")


if __name__ == "__main__":
    run()
