"""新品查重 Web 服务"""
import sys, os, io, threading, time, webbrowser, requests

if getattr(sys, 'frozen', False):
    BASE = sys._MEIPASS
    EXE_DIR = os.path.dirname(sys.executable)
    sys.path.insert(0, os.path.join(BASE, 'scripts'))
    TEMPLATES = os.path.join(BASE, 'web', 'templates')
    INDEX_PATH = os.path.join(EXE_DIR, 'index.json')
else:
    BASE = os.path.dirname(os.path.abspath(__file__))
    EXE_DIR = BASE
    sys.path.insert(0, os.path.join(BASE, '..', 'scripts'))
    TEMPLATES = os.path.join(BASE, 'templates')
    INDEX_PATH = os.path.join(BASE, '..', 'index.json')

from flask import Flask, render_template, request, jsonify, send_file
from translator import SemanticTranslator, ProductRow, resolve_spec
from matcher import load_index, check_one, IndexedRow
import openpyxl

app = Flask(__name__, template_folder=TEMPLATES)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB
translator = SemanticTranslator()
try:
    index = load_index(INDEX_PATH)
except FileNotFoundError:
    index = {}

COLS = {'brand': 2, 'name': 3, 'spec': 4, 'category': 5, 'unit': 6, 'remark': 7}

task_state = {
    "running": False, "paused": False, "done": 0, "total": 0,
    "results": [], "wb": None,
    "api_key": "", "model": "deepseek-chat", "api_url": "https://api.deepseek.com/v1/chat/completions",
    "clean_progress": {"done": 0, "total": 0, "running": False},
    "clean_results": [],
    "build_progress": {"done": 0, "total": 0, "running": False},
    "index": None,
    "review_progress": {"done": 0, "total": 0},
}


@app.route('/')
def home():
    print("[API] / 被调用")
    return render_template('index.html')


@app.route('/api/test-connection', methods=['POST'])
def test_connection():
    print("[API] /api/test-connection")
    data = request.get_json() or {}
    api_key = data.get('api_key', task_state['api_key'])
    api_url = data.get('api_url', task_state['api_url'])
    model = data.get('model', task_state['model'])
    if not api_key:
        return jsonify({'ok': False, 'error': '未提供 API Key'})
    try:
        resp = requests.post(api_url,
            headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
            json={'model': model, 'messages': [{'role': 'user', 'content': 'hi'}], 'max_tokens': 5},
            timeout=10)
        if resp.status_code == 200 and resp.json().get('choices'):
            return jsonify({'ok': True})
        return jsonify({'ok': False, 'error': f'状态码{resp.status_code}: {resp.text[:100]}'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/settings', methods=['POST'])
def save_settings():
    print("[API] /api/settings")
    data = request.get_json()
    if not data:
        return jsonify({'error': '无效请求'}), 400
    task_state['api_key'] = data.get('api_key', task_state['api_key'])
    task_state['model'] = data.get('model', task_state['model'])
    task_state['api_url'] = data.get('api_url', task_state['api_url'])
    import judge
    judge.API_KEY = task_state['api_key']
    if task_state['api_url'] != 'https://api.deepseek.com/v1':
        judge.API_URL = task_state['api_url']
    return jsonify({'ok': True, 'has_key': bool(task_state['api_key'])})


@app.route('/api/upload', methods=['POST'])
def upload():
    print("[API] /api/upload")
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请上传文件'}), 400
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '请上传 .xlsx 文件'}), 400
    try:
        wb = openpyxl.load_workbook(file)
    except Exception:
        return jsonify({'error': '无法读取文件，请确认格式正确'}), 400
    ws = wb['工作表1'] if '工作表1' in wb.sheetnames else wb.active
    # 检查名称列
    has_name = False
    for r in range(2, min(ws.max_row + 1, 5)):
        if str(ws.cell(r, COLS['name']).value or '').strip():
            has_name = True; break
    if not has_name:
        return jsonify({'error': '表格缺少*名称列（第3列）'}), 400
    preview = []
    total = 0
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, COLS['name']).value or '').strip()
        if not name: continue
        total += 1
        if len(preview) < 20:
            preview.append({'row': r, 'name': name,
                'brand': str(ws.cell(r, COLS['brand']).value or '').strip(),
                'spec': str(ws.cell(r, COLS['spec']).value or '').strip(),
                'unit': str(ws.cell(r, COLS['unit']).value or '').strip()})
    task_state['wb'] = wb
    task_state['total'] = total
    return jsonify({'preview': preview, 'total': total})


@app.route('/api/start', methods=['POST'])
def start():
    print("[API] /api/start")
    if not task_state.get('index'):
        print("[API] /api/start 被拒绝: 索引未建立")
        return jsonify({'error': '请先在Step1上传商品库建立索引'}), 400
    if task_state['running']:
        return jsonify({'error': '已有任务在运行'}), 400
    if not task_state['api_key']:
        return jsonify({'error': '请先在设置中填写 API Key'}), 400
    if not task_state['wb']:
        return jsonify({'error': '请先上传接龙表'}), 400
    task_state['running'] = True
    task_state['paused'] = False
    task_state['done'] = 0
    task_state['results'] = []
    threading.Thread(target=process, daemon=True).start()
    return jsonify({'ok': True, 'total': task_state['total']})


def process():
    ws = task_state['wb']['工作表1'] if '工作表1' in task_state['wb'].sheetnames else task_state['wb'].active
    for r in range(2, ws.max_row + 1):
        if not task_state['running']: break
        while task_state['paused']:
            time.sleep(0.5)
            if not task_state['running']: return
        name = str(ws.cell(r, COLS['name']).value or '').strip()
        if not name: continue
        brand_raw = str(ws.cell(r, COLS['brand']).value or '').strip()
        brand = None if (not brand_raw or brand_raw in ('None', '无要求')) else brand_raw
        try:
            result = check_one(translator, name, brand,
                str(ws.cell(r, COLS['spec']).value or '').strip(),
                str(ws.cell(r, COLS['unit']).value or '').strip(),
                str(ws.cell(r, COLS['category']).value or '').strip(),
                str(ws.cell(r, COLS['remark']).value or '').strip(),
                task_state['index'])
            if result.result == '建议复用' and result.suggested_spuid:
                ws.cell(r, 12, value=result.suggested_spuid)
            ws.cell(r, 13, value=result.result)
            ws.cell(r, 14, value=result.detail)
            task_state['results'].append({'row': r, 'name': name, 'unit': str(ws.cell(r, COLS['unit']).value or ''),
                'result': result.result, 'spuid': result.suggested_spuid or '', 'detail': result.detail,
                'brand': brand or '', 'spec': str(ws.cell(r, COLS['spec']).value or '').strip(),
                'cat': str(ws.cell(r, COLS['category']).value or '').strip()})
            print(f"[查重] {task_state['done']}/{task_state['total']} {name[:20]} → {result.result}")
        except Exception as e:
            task_state['results'].append({'row': r, 'name': name, 'unit': '',
                'result': '待确认', 'spuid': '', 'detail': f'异常:{e}'})
        task_state['done'] += 1
    task_state['running'] = False


@app.route('/api/progress')
def progress():
    print("[API] /api/progress")
    return jsonify({'done': task_state['done'], 'total': task_state['total'],
                    'running': task_state['running'], 'paused': task_state['paused']})


@app.route('/api/results')
def results():
    print("[API] /api/results")
    return jsonify({'results': task_state['results'], 'total': len(task_state['results'])})


@app.route('/api/stop', methods=['POST'])
def stop():
    print("[API] /api/stop")
    task_state['running'] = False
    task_state['paused'] = False
    return jsonify({'ok': True})


@app.route('/api/pause', methods=['POST'])
def pause():
    print("[API] /api/pause")
    task_state['paused'] = not task_state['paused']
    return jsonify({'ok': True, 'paused': task_state['paused']})


@app.route('/api/export')
def export():
    print("[API] /api/export")
    if not task_state['wb']:
        return jsonify({'error': '无结果可导出'}), 400
    output = io.BytesIO()
    task_state['wb'].save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='查重结果.xlsx')


def build_process(wb):
    ws = wb['工作表1'] if '工作表1' in wb.sheetnames else wb.active
    idx = {}
    total = 0
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 2).value or '').strip()
        if not name: continue
        total += 1
        nm = translator.translate_name(name, str(ws.cell(r, 5).value or ''))
        dm = translator.translate_desc(str(ws.cell(r, 4).value or ''), str(ws.cell(r, 3).value or ''))
        spec = resolve_spec(nm, dm, False)
        row = IndexedRow(spuid=str(ws.cell(r, 1).value or ''), core=nm.core,
            brand=nm.brand, spec_core=spec.spec_core, has_processing=False,
            unit=str(ws.cell(r, 3).value or ''), raw_name=str(ws.cell(r, 2).value or ''),
            category=str(ws.cell(r, 5).value or ''), has_chaoma=nm.has_chaoma,
            has_status=nm.has_status, hit_count=0)
        idx.setdefault(nm.core, []).append(row)
        task_state['build_progress']['done'] = total
        if total % 200 == 0:
            print(f"[建索引] {total}")
    task_state['index'] = idx
    task_state['build_progress']['total'] = total
    task_state['build_progress']['running'] = False
    print(f"[建索引] 完成 {total} 条, {len(idx)} 个core")


def clean_process(wb):
    from rules import RuleEngine
    ws = wb['工作表1'] if '工作表1' in wb.sheetnames else wb.active
    rows = []
    total = 0
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 2).value or '').strip()
        if not name: continue
        total += 1
        nm = translator.translate_name(name, str(ws.cell(r, 5).value or ''))
        dm = translator.translate_desc(str(ws.cell(r, 4).value or ''), str(ws.cell(r, 3).value or ''))
        rows.append(ProductRow(spuid=str(ws.cell(r, 1).value or ''), name_meaning=nm,
            unit=str(ws.cell(r, 3).value or ''), desc_meaning=dm,
            category=str(ws.cell(r, 5).value or '')))
    task_state['clean_progress']['total'] = total
    engine = RuleEngine(rows, translator)
    engine.run_all()
    for row in rows:
        task_state['clean_results'].append({
            'spuid': row.spuid, 'name': row.name_meaning.raw, 'unit': row.unit,
            'desc': row.desc_meaning.raw, 'category': row.category,
            'group_id': row.group_id or '', 'group_desc': row.group_desc or '',
            'suggestion': row.suggestion or '', 'anomaly_class': row.anomaly_class or ''})
        task_state['clean_progress']['done'] += 1
        if task_state['clean_progress']['done'] % 200 == 0:
            print(f"[清洗] {task_state['clean_progress']['done']}/{total}")
    # === DeepSeek 复核 ===
    if not task_state.get('api_key'):
        print("[复核] 未设置 API Key，跳过复核")
        task_state['clean_progress']['running'] = False
        return

    import judge
    judge.API_KEY = task_state['api_key']
    judge.API_URL = task_state.get('api_url', 'https://api.deepseek.com/v1/chat/completions')
    model = task_state.get('model', 'deepseek-chat')

    review_classes = {'完全重复', '抄码名重复', '异名同物'}
    candidates = [r for r in task_state['clean_results'] if r['anomaly_class'] in review_classes]
    if not candidates:
        print("[复核] 无需要复核的行")
        task_state['clean_progress']['running'] = False
        return

    from collections import defaultdict
    import re
    groups_dict = defaultdict(list)
    for r in candidates:
        gid = r['group_id']
        if gid:
            groups_dict[gid].append(r)

    # 构造批量组列表
    group_list = []
    for gid, items in groups_dict.items():
        review_items = []
        for r in items:
            clean_name = re.sub(r'[-][A-Z]{2}\d+[【〔]?.*', '', r['name']).strip()
            review_items.append({'name': clean_name, 'unit': r['unit'],
                                 'desc': r['desc'], 'category': r['category']})
        group_list.append({'group_id': gid, 'items': review_items})

    total_groups = len(group_list)
    task_state['review_progress'] = {'done': 0, 'total': total_groups}
    print(f"[复核] 开始, 共 {total_groups} 组, 批量 {10} 组/次")

    yes_count = no_count = err_count = 0
    BATCH_SIZE = 10

    for batch_start in range(0, total_groups, BATCH_SIZE):
        batch = group_list[batch_start:batch_start + BATCH_SIZE]
        results = judge.review_batch(batch, model)

        for gid, result, reason in results:
            items = groups_dict[gid]
            if result is True:
                yes_count += 1
                print(f"[复核] {gid} → yes")
            elif result is False:
                no_count += 1
                print(f"[复核] {gid} → no")
                for r in items:
                    r['anomaly_class'] = '正常'
                    r['group_id'] = ''
                    r['suggestion'] = ''
            else:
                err_count += 1
                print(f"[复核] {gid} → 异常: {reason}")
                for r in items:
                    r['anomaly_class'] = '待确认'
                    r['suggestion'] = reason

        task_state['review_progress']['done'] = min(batch_start + BATCH_SIZE, total_groups)
        print(f"[复核] 进度 {task_state['review_progress']['done']}/{total_groups}")

    print(f"[复核] 完成: {total_groups}组, yes={yes_count}, no={no_count}, 异常={err_count}")
    task_state['clean_progress']['running'] = False


@app.route('/api/clean', methods=['POST'])
def clean():
    print("[API] /api/clean")
    if task_state['clean_progress']['running']:
        return jsonify({'error': '已有清洗任务在运行'}), 400
    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        return jsonify({'error': '请上传 .xlsx 文件'}), 400
    try:
        wb = openpyxl.load_workbook(file)
    except Exception:
        return jsonify({'error': '无法读取文件，请确认格式正确'}), 400
    task_state['clean_progress'] = {'done': 0, 'total': 0, 'running': True}
    task_state['clean_results'] = []
    threading.Thread(target=clean_process, args=(wb,), daemon=True).start()
    return jsonify({'ok': True})


@app.route('/api/clean-progress')
def clean_progress():
    print("[API] /api/clean-progress")
    p = dict(task_state['clean_progress'])
    p['review'] = dict(task_state.get('review_progress', {'done': 0, 'total': 0}))
    return jsonify(p)


@app.route('/api/clean-export')
def clean_export():
    print("[API] /api/clean-export")
    if not task_state['clean_results']:
        return jsonify({'error': '无清洗结果可导出'}), 400
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '工作表1'
    ws.append(['SPUID', '名称', '单位', '描述', '异常分类', '组号', '组说明', '建议'])
    for row in task_state['clean_results']:
        ws.append([row['spuid'], row['name'], row['unit'], row['desc'],
                   row['anomaly_class'], row['group_id'], row['group_desc'],
                   row['suggestion']])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='清洗结果.xlsx')


@app.route('/api/clean-results')
def clean_results():
    print("[API] /api/clean-results")
    return jsonify({'results': task_state['clean_results'],
                    'total': len(task_state['clean_results'])})


@app.route('/api/build-index', methods=['POST'])
def build_index():
    print("[API] /api/build-index")
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请上传文件'}), 400
    try:
        wb = openpyxl.load_workbook(file)
    except Exception:
        return jsonify({'error': '无法读取文件，请确认格式正确'}), 400
    task_state['build_progress'] = {'done': 0, 'total': 0, 'running': True}
    threading.Thread(target=build_process, args=(wb,), daemon=True).start()
    return jsonify({'ok': True})


@app.route('/api/build-progress')
def build_progress():
    print("[API] /api/build-progress")
    return jsonify(task_state['build_progress'])


def main():
    port = int(os.environ.get('WEB_PORT', 5000))
    url = f'http://127.0.0.1:{port}'
    webbrowser.open(url)
    print(f'服务已启动: {url}')
    app.run(host='127.0.0.1', port=port, debug=False)


if __name__ == '__main__':
    main()
